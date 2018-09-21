import win32com.client
import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl


def ensure_dir(file_path):
    if not os.path.exists(file_path):
        os.makedirs(file_path)


def print_data(m):
    date = m.SentOn.strftime("%d-%m-%y")
    sender = m.Sender().replace ('|','-').replace (':','-').replace ('/','')
    #print(date)
    #print(m)
    index = 0
    pdf_attachement = ''
    att_filename = None
    for att in m.Attachments:
        ensure_dir("D:\PycharmProjects\invoice_printer\invoices\\" + sender)
        att_filename ="D:\PycharmProjects\invoice_printer\invoices\\" + sender +"\\" + date + "-"+str(att)
        try:
            att.SaveAsFile(att_filename)
        except:
            pass
        if '.pdf' in att_filename:
            pdf_attachement = att_filename
    if pdf_attachement == '':
        pdf_attachement = att_filename
    return [date,sender,m.SenderEmailAddress,m.Subject,pdf_attachement]

def get_distribution(messages, num_items):
    keywords = ['distribution report']
    index = 0
    data = []
    for m in messages:
        try:
            for k in keywords:
                if k in m.Subject.lower():
                    data.append(print_data(m))
                    #print(m, m.Attachments[0])
                    break
        except:
            pass
        if index > num_items:
            break
        index += 1
    
    return data


def get_invoices(messages, num_items):
    keywords = ['invoice',"has been confirmed",'aviv packing house to yyz', 'Prebook #']
    index = 0
    data = []
    for m in messages:
        try:
            for k in keywords:
                if k in m.Subject.lower():
                    data.append(print_data(m))
                    break
        except Exception as e:
            print(e)
            print(m.sender())
            print(m)
            for att in m.Attachments:
                print(att)
        if index > num_items:
            break
        index += 1
    return data

def add_checked(filename, headings, data):
    wb2 = load_workbook(filename)
    ws2 = wb2.active

    index = 0
    checked = []
    for row in ws2.rows:
        if index == 0:
            index += 1
            categories_order = row
        else:
            p = {}
            for i in range(len(categories_order)):
                p[categories_order[i].value] = row[i].value
            
            checked.append(p)

    for c in checked:
        t  = []
        for h in headings:
            t.append(c[h])
        
        for d in data:
            
            if t[:-2] == d:
                d.extend(t[-2:])
                break

def update_data(num_msg = 1000):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # "6" refers to the index of a folder - in this case,
    # the inbox. You can change that number to reference
    # any other folder

    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)

    return get_invoices(messages, num_msg)

def update(num_msg = 1000):
    filename = 'invoice_list.xlsx'
    headings = ["Date","Sender", "Email","Subject","filename", "printed", "signed"]
    data = update_data(num_msg)

    add_checked(filename, headings, data)
    #data.sort()
    #data.reverse()

    wb = Workbook()
    ws = wb.active

    ws.append(headings)
    for d in data:
        ws.append(d)


    tab = Table(displayName="Invoices", ref="A1:G%s" % (len(data) +1))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
        
    wb.save(filename)
    os.startfile(filename)


filename = r'invoice_list.xlsx'
headings = ["Date","Sender", "Email","Subject","filename", "printed", "signed"]

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6) # "6" refers to the index of a folder - in this case,
                                    # the inbox. You can change that number to reference
                                    # any other folder
                                    
messages = inbox.Items
messages.Sort("[ReceivedTime]", True)



if __name__ == "__main__":

    data = get_invoices(messages, 4000)
    add_checked(filename, headings, data)
    #data.sort()
    #data.reverse()

    wb = Workbook()
    ws = wb.active

    ws.append(headings)

    for d in data:
        ws.append(d)


    tab = Table(displayName="Invoices", ref="A1:G%s" % len(data))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
        
    try:
        wb.save(filename)
    except(PermissionError):
        print("Already Open")
    os.startfile(filename)

