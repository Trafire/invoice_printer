from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import time
import print_report

def excel_to_dict(ws):
    index = 0
    data = []
    for row in ws:
        if index == 0:
            headings = []
            for cell in row:
                headings.append(cell.value)
            index += 1
        else:
            col = 0
            r = {}
            for cell in headings:
                r[headings[col]] = str(row[col].value).strip()
                col +=1
            if r not in data:
                data.append(r)
    return data

loop = True
while loop:
    excel_filename = "print_list2.xlsx"
    wb = load_workbook(excel_filename)
    ws= wb.active

    data = excel_to_dict(ws)
    for d in data:
        if not d['invoice_num'] or not d['date']or not d['supplier']:
            os.startfile(d['filename'])
            input("next?")
            break
        else:
            loop = False
            break
ran = []

for d in data:
     date = d['date']
     supplier = d['supplier']
     invoice_num = d['invoice_num']
     filename = d['filename']
     if True: #(date ,supplier) not in ran:
         if d['completed'] in  ('None', ''):
             print_report.get_report(d['date'], d['supplier'], d['filename'], d['invoice_num'])
             ran.append((date, supplier))
             time.sleep(2)
     else:
         print_report.update_data(date, supplier, invoice_num, filename)


time.sleep(8)


for d in data:
    if d['completed'] == 'None':
        d['completed'] = ''
        try:
            if print_report.print_total(d['date'], d['supplier'], d['filename']):
                    d['completed'] = 'x'
        except(KeyError):
            print(d)



headings = list(data[0].keys())
wb2 = Workbook()
ws2 = wb2.active
ws2.append(headings)
for d in data:
    row = []
    for h in headings:
        row.append(d[h])
    ws2.append(row)
try:
    wb2.save(excel_filename)
except(PermissionError):

    count = 1
    excel_filename = r"print_list%s.xlsx" % count
    while os.path.exists(excel_filename):
        count += 1
        excel_filename = r"print_list%s.xlsx" % count
    wb2.save(excel_filename)

os.startfile(excel_filename)